# -*- coding: utf-8 -*-
"""Generate SliceMatic_Calculator.xlsx — a live, plug-and-play unit-economics model.

Every metric is a real Excel formula referencing the Inputs sheet, so changing any
yellow input cell recalculates the whole workbook.
"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ---- styling helpers ----------------------------------------------------------
INPUT_FILL = PatternFill("solid", fgColor="FFF2CC")    # yellow = editable
HEAD_FILL  = PatternFill("solid", fgColor="C0392B")     # red banner
SUB_FILL   = PatternFill("solid", fgColor="F2DEDE")     # light section header
CALC_FILL  = PatternFill("solid", fgColor="EAF2F8")     # blue-ish = computed result
TOTAL_FILL = PatternFill("solid", fgColor="D5E8D4")     # green = key total

WHITE_BOLD = Font(bold=True, color="FFFFFF", size=12)
BOLD       = Font(bold=True)
ITALIC     = Font(italic=True, color="666666", size=9)
TITLE_FONT = Font(bold=True, size=14, color="C0392B")

thin = Side(style="thin", color="CCCCCC")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)

RUPEE = u'₹#,##0.00'
RUPEE0 = u'₹#,##0'
PCT = '0.0%'
NUM1 = '#,##0.0'
NUM0 = '#,##0'


def banner(ws, row, text, span=3):
    c = ws.cell(row=row, column=1, value=text)
    c.font = WHITE_BOLD
    c.fill = HEAD_FILL
    for col in range(1, span + 1):
        ws.cell(row=row, column=col).fill = HEAD_FILL
    ws.row_dimensions[row].height = 20


def section(ws, row, text, span=3):
    c = ws.cell(row=row, column=1, value=text)
    c.font = BOLD
    for col in range(1, span + 1):
        ws.cell(row=row, column=col).fill = SUB_FILL


def label_input(ws, row, label, value, fmt=RUPEE0, note=""):
    ws.cell(row=row, column=1, value=label)
    c = ws.cell(row=row, column=2, value=value)
    c.fill = INPUT_FILL
    c.number_format = fmt
    c.border = BORDER
    if note:
        n = ws.cell(row=row, column=3, value=note)
        n.font = ITALIC


def label_calc(ws, row, label, formula, fmt=RUPEE0, note="", key=False):
    lc = ws.cell(row=row, column=1, value=label)
    c = ws.cell(row=row, column=2, value=formula)
    c.number_format = fmt
    c.fill = TOTAL_FILL if key else CALC_FILL
    c.border = BORDER
    if key:
        lc.font = BOLD
        c.font = BOLD
    if note:
        n = ws.cell(row=row, column=3, value=note)
        n.font = ITALIC


wb = Workbook()

# ==============================================================================
# SHEET 1 — INPUTS
# ==============================================================================
ws = wb.active
ws.title = "Inputs"
ws.column_dimensions['A'].width = 34
ws.column_dimensions['B'].width = 16
ws.column_dimensions['C'].width = 50

ws['A1'] = "SliceMatic — Inputs (edit the yellow cells only)"
ws['A1'].font = TITLE_FONT

banner(ws, 3, "FIXED COSTS  (Rs. / month)")
fixed_items = [
    ("Kitchen rent", 55000, "1,200 sq ft commercial lease"),
    ("Equipment EMI", 14500, "Oven, mixer, refrigeration — 36-mo loan"),
    ("Electricity", 12600, "~1,800 units/month"),
    ("Gas / LPG", 5550, "3 cylinders @ Rs.1,850"),
    ("Head chef", 28000, ""),
    ("Kitchen helper", 16500, "min wage + ESI/PF"),
    ("Counter + billing staff", 18000, ""),
    ("Delivery riders (2)", 32000, "fixed component only"),
    ("Internet + POS + SaaS", 2800, ""),
    ("Marketing (fixed)", 8000, ""),
    ("Packaging (fixed component)", 4200, ""),
    ("Misc / contingency", 5760, "3% maintenance/consumables"),
]
r = 4
for name, val, note in fixed_items:
    label_input(ws, r, name, val, RUPEE0, note)
    r += 1
TOTAL_FIXED_ROW = r
label_calc(ws, r, "TOTAL FIXED COSTS", f"=SUM(B4:B{r-1})", RUPEE0, "auto-sum", key=True)
r += 2

banner(ws, r, "REVENUE DRIVERS")
rev_start = r + 1
WD_DAYS = r + 1; label_input(ws, WD_DAYS, "Weekday days / month", 22, NUM0)
WE_DAYS = r + 2; label_input(ws, WE_DAYS, "Weekend+holiday days / month", 8, NUM0)
WD_ORD  = r + 3; label_input(ws, WD_ORD, "Weekday orders / day", 38, NUM0)
WE_ORD  = r + 4; label_input(ws, WE_ORD, "Weekend orders / day", 68, NUM0)
WD_AOV  = r + 5; label_input(ws, WD_AOV, "Weekday AOV (ex-GST)", 792, RUPEE0)
WE_AOV  = r + 6; label_input(ws, WE_AOV, "Weekend AOV (ex-GST)", 940, RUPEE0)
r = r + 8

banner(ws, r, "VARIABLE COST PER ORDER")
ING   = r + 1; label_input(ws, ING, "Ingredient COGS / order", 148, RUPEE0)
PKG   = r + 2; label_input(ws, PKG, "Packaging / order", 18, RUPEE0)
DELV  = r + 3; label_input(ws, DELV, "Delivery incentive / order", 22, RUPEE0)
GWPCT = r + 4; label_input(ws, GWPCT, "Payment gateway %", 0.018, PCT, "% of order value")
r = r + 6

banner(ws, r, "PRICING RULES")
GST   = r + 1; label_input(ws, GST, "GST rate", 0.18, PCT)
DTHR  = r + 2; label_input(ws, DTHR, "Discount threshold (qty >=)", 5, NUM0)
DPCT  = r + 3; label_input(ws, DPCT, "Discount %", 0.10, PCT)
r = r + 5

banner(ws, r, "CAPACITY  &  INVESTMENT")
MAXDAY = r + 1; label_input(ws, MAXDAY, "Max orders / day (100% capacity)", 80, NUM0)
UTIL   = r + 2; label_input(ws, UTIL, "Planned utilisation %", 0.60, PCT)
INVEST = r + 3; label_input(ws, INVEST, "Initial investment", 1500000, RUPEE0)

# named cell references for use elsewhere
I = "Inputs!"
F_FIXED = f"{I}B{TOTAL_FIXED_ROW}"
F_RENT  = f"{I}B4"
F_WDDAYS, F_WEDAYS = f"{I}B{WD_DAYS}", f"{I}B{WE_DAYS}"
F_WDORD, F_WEORD = f"{I}B{WD_ORD}", f"{I}B{WE_ORD}"
F_WDAOV, F_WEAOV = f"{I}B{WD_AOV}", f"{I}B{WE_AOV}"
F_ING, F_PKG, F_DELV, F_GW = f"{I}B{ING}", f"{I}B{PKG}", f"{I}B{DELV}", f"{I}B{GWPCT}"
F_GST, F_DTHR, F_DPCT = f"{I}B{GST}", f"{I}B{DTHR}", f"{I}B{DPCT}"
F_MAXDAY, F_UTIL, F_INVEST = f"{I}B{MAXDAY}", f"{I}B{UTIL}", f"{I}B{INVEST}"

# ==============================================================================
# SHEET 2 — DASHBOARD
# ==============================================================================
d = wb.create_sheet("Dashboard")
d.column_dimensions['A'].width = 34
d.column_dimensions['B'].width = 18
d.column_dimensions['C'].width = 40

d['A1'] = "SliceMatic — Live Metrics Dashboard"
d['A1'].font = TITLE_FONT
d['A2'] = "All values computed from the Inputs sheet. Edit inputs there; these update automatically."
d['A2'].font = ITALIC

banner(d, 4, "MONTHLY VOLUME & REVENUE")
MORD = 5;  label_calc(d, MORD, "Monthly orders", f"={F_WDDAYS}*{F_WDORD}+{F_WEDAYS}*{F_WEORD}", NUM0)
ADAY = 6;  label_calc(d, ADAY, "Avg orders / day", f"=B{MORD}/({F_WDDAYS}+{F_WEDAYS})", NUM1)
MREV = 7;  label_calc(d, MREV, "Monthly revenue (ex-GST)",
                      f"={F_WDDAYS}*{F_WDORD}*{F_WDAOV}+{F_WEDAYS}*{F_WEORD}*{F_WEAOV}", RUPEE0)
BAOV = 8;  label_calc(d, BAOV, "Blended AOV (ex-GST)", f"=B{MREV}/B{MORD}", RUPEE0)
GSTC = 9;  label_calc(d, GSTC, "GST collected (passed to Govt.)", f"=B{MREV}*{F_GST}", RUPEE0)
GROSS= 10; label_calc(d, GROSS,"Gross billed (incl GST)", f"=B{MREV}+B{GSTC}", RUPEE0)

banner(d, 12, "PER-ORDER ECONOMICS")
GWF = 13; label_calc(d, GWF, "Gateway fee / order", f"=B{BAOV}*{F_GW}", RUPEE)
CMO = 14; label_calc(d, CMO, "Contribution margin / order",
                     f"=B{BAOV}-{F_ING}-{F_PKG}-{F_DELV}-B{GWF}", RUPEE, key=True)
CMP = 15; label_calc(d, CMP, "Contribution margin %", f"=B{CMO}/B{BAOV}", PCT)

banner(d, 17, "MONTHLY P&L")
MCON = 18; label_calc(d, MCON, "Monthly contribution", f"=B{MORD}*B{CMO}", RUPEE0)
MFIX = 19; label_calc(d, MFIX, "Less: total fixed costs", f"=-{F_FIXED}", RUPEE0)
EBIT = 20; label_calc(d, EBIT, "EBITDA (operating profit)", f"=B{MCON}+B{MFIX}", RUPEE0, key=True)
OPM  = 21; label_calc(d, OPM,  "Operating margin", f"=B{EBIT}/B{MREV}", PCT)

banner(d, 23, "BREAK-EVEN")
BEM = 24; label_calc(d, BEM, "Break-even orders / month", f"={F_FIXED}/B{CMO}", NUM0)
BED = 25; label_calc(d, BED, "Break-even orders / day", f"=B{BEM}/30", NUM1, key=True)
SAF = 26; label_calc(d, SAF, "Safety multiple (plan / break-even)", f"=B{ADAY}/B{BED}", '0.0"x"')
CAPU= 27; label_calc(d, CAPU,"Capacity utilisation (implied)", f"=B{ADAY}/{F_MAXDAY}", PCT)

banner(d, 29, "PAYBACK")
PBK = 30; label_calc(d, PBK, "Simple payback (months)", f"={F_INVEST}/B{EBIT}", NUM1,
                     "at current EBITDA; real ramp is slower", key=True)

# dashboard refs
D = "Dashboard!"
D_MORD, D_BAOV, D_CMO, D_EBIT, D_MREV = f"{D}B{MORD}", f"{D}B{BAOV}", f"{D}B{CMO}", f"{D}B{EBIT}", f"{D}B{MREV}"

# ==============================================================================
# SHEET 3 — COGS & MARGIN
# ==============================================================================
c = wb.create_sheet("COGS & Margin")
for col, w in zip("ABCDE", [28, 16, 16, 16, 16]):
    c.column_dimensions[col].width = w
c['A1'] = "COGS & Gross Margin by Pizza Type"
c['A1'].font = TITLE_FONT
c['A2'] = "Edit yellow component costs and selling prices; totals & margins recalc."
c['A2'].font = ITALIC

hdr = ["Component", "Margherita", "Farm House", "Cheese Burst", "Custom"]
for j, h in enumerate(hdr, start=1):
    cell = c.cell(row=4, column=j, value=h)
    cell.font = WHITE_BOLD; cell.fill = HEAD_FILL

rows = [
    ("Pizza base ingredients", 38, 45, 72, 50),
    ("Sauce + seasoning", 12, 14, 14, 13),
    ("Cheese (mozzarella)", 28, 32, 65, 38),
    ("Pizza-specific toppings", 8, 22, 18, 17),
    ("Add-on topping (avg 1)", 10, 10, 10, 10),
    ("Packaging (box + bag)", 18, 18, 18, 18),
    ("Delivery variable", 22, 22, 22, 22),
]
start = 5
for i, (name, *vals) in enumerate(rows):
    rr = start + i
    c.cell(row=rr, column=1, value=name)
    for j, v in enumerate(vals, start=2):
        cell = c.cell(row=rr, column=j, value=v)
        cell.fill = INPUT_FILL; cell.number_format = RUPEE0; cell.border = BORDER
last = start + len(rows) - 1
tot = last + 1
c.cell(row=tot, column=1, value="TOTAL COGS / pizza").font = BOLD
for j in range(2, 6):
    col = get_column_letter(j)
    cell = c.cell(row=tot, column=j, value=f"=SUM({col}{start}:{col}{last})")
    cell.number_format = RUPEE0; cell.fill = TOTAL_FILL; cell.font = BOLD; cell.border = BORDER

sp = tot + 1
c.cell(row=sp, column=1, value="Selling price (base+pizza+1 topping)").font = BOLD
for j, v in zip(range(2, 6), [397, 417, 447, 420]):
    cell = c.cell(row=sp, column=j, value=v)
    cell.fill = INPUT_FILL; cell.number_format = RUPEE0; cell.border = BORDER

gm = sp + 1
c.cell(row=gm, column=1, value="Gross margin / pizza").font = BOLD
for j in range(2, 6):
    col = get_column_letter(j)
    cell = c.cell(row=gm, column=j, value=f"={col}{sp}-{col}{tot}")
    cell.number_format = RUPEE0; cell.fill = CALC_FILL; cell.border = BORDER

gmp = gm + 1
c.cell(row=gmp, column=1, value="Gross margin %").font = BOLD
for j in range(2, 6):
    col = get_column_letter(j)
    cell = c.cell(row=gmp, column=j, value=f"={col}{gm}/{col}{sp}")
    cell.number_format = PCT; cell.fill = CALC_FILL; cell.border = BORDER

CB_TOT = f"'COGS & Margin'!D{tot}"   # Cheese Burst total COGS (for Q4)

# ==============================================================================
# SHEET 4 — BILL CALCULATOR
# ==============================================================================
b = wb.create_sheet("Bill Calculator")
b.column_dimensions['A'].width = 32
b.column_dimensions['B'].width = 16
b.column_dimensions['C'].width = 40
b['A1'] = "Single-Order Bill Calculator"
b['A1'].font = TITLE_FONT
b['A2'] = "Plug in prices + quantity (yellow). Mirrors the app's pricing logic exactly."
b['A2'].font = ITALIC

banner(b, 4, "ORDER INPUTS")
BP = 5; label_input(b, BP, "Base price", 229, RUPEE0)
PP = 6; label_input(b, PP, "Pizza price", 379, RUPEE0)
TP = 7; label_input(b, TP, "Topping price", 69, RUPEE0)
QTY= 8; label_input(b, QTY,"Quantity (1-10)", 5, NUM0)

banner(b, 10, "COMPUTED BILL")
UNIT= 11; label_calc(b, UNIT,"Unit price (base+pizza+topping)", f"=B{BP}+B{PP}+B{TP}", RUPEE)
SUB = 12; label_calc(b, SUB, "Subtotal (unit x qty)", f"=B{UNIT}*B{QTY}", RUPEE)
DISC= 13; label_calc(b, DISC,"Discount", f"=IF(B{QTY}>={F_DTHR},B{SUB}*{F_DPCT},0)", RUPEE)
PD  = 14; label_calc(b, PD,  "Post-discount subtotal", f"=B{SUB}-B{DISC}", RUPEE)
GS  = 15; label_calc(b, GS,  "GST", f"=B{PD}*{F_GST}", RUPEE)
TOT = 16; label_calc(b, TOT, "TOTAL PAYABLE", f"=B{PD}+B{GS}", RUPEE, key=True)

# ==============================================================================
# SHEET 5 — SCENARIOS (the six stress tests)
# ==============================================================================
s = wb.create_sheet("Scenarios")
s.column_dimensions['A'].width = 40
s.column_dimensions['B'].width = 16
s.column_dimensions['C'].width = 44

s['A1'] = "Stress Tests — Challenge These Numbers"
s['A1'].font = TITLE_FONT
s['A2'] = "Yellow = scenario inputs. Everything else recalculates from Inputs + Dashboard."
s['A2'].font = ITALIC

# ---- Q1: rent shock ----
banner(s, 4, "Q1  —  RENT SHOCK")
NRENT = 5; label_input(s, NRENT, "New monthly rent", 70000, RUPEE0)
NFIX  = 6; label_calc(s, NFIX, "New total fixed costs", f"={F_FIXED}-{F_RENT}+B{NRENT}", RUPEE0)
NBEM  = 7; label_calc(s, NBEM, "New break-even orders / month", f"=B{NFIX}/{D_CMO}", NUM0)
NBED  = 8; label_calc(s, NBED, "New break-even orders / day", f"=B{NBEM}/30", NUM1, key=True)
UNV   = 9; label_calc(s, UNV, "Rent that makes plan unviable",
                      f"={D_MORD}*{D_CMO}-({F_FIXED}-{F_RENT})", RUPEE0,
                      "rent where break-even = planned volume", key=True)

# ---- Q2: aggregator ----
banner(s, 11, "Q2  —  AGGREGATOR (Zomato/Swiggy)")
ASH = 12; label_input(s, ASH, "Aggregator order share", 0.40, PCT)
ACM = 13; label_input(s, ACM, "Aggregator commission %", 0.25, PCT)
AGGCM = 14; label_calc(s, AGGCM, "CM / aggregator order",
                       f"={D_BAOV}-{F_ING}-{F_PKG}-{D_BAOV}*B{ACM}", RUPEE,
                       "aggregator covers delivery + payment")
DIRCM = 15; label_calc(s, DIRCM, "CM / direct order", f"={D_CMO}", RUPEE)
BLCM  = 16; label_calc(s, BLCM, "Blended CM / order",
                       f"=B{ASH}*B{AGGCM}+(1-B{ASH})*B{DIRCM}", RUPEE, key=True)
ABE   = 17; label_calc(s, ABE, "New break-even orders / day",
                       f"={F_FIXED}/B{BLCM}/30", NUM1, key=True)

# ---- Q3: 3rd rider ----
banner(s, 19, "Q3  —  3rd RIDER JUSTIFICATION")
RCOST = 20; label_input(s, RCOST, "3rd rider cost / month", 16000, RUPEE0)
RTRIP = 21; label_input(s, RTRIP, "Deliveries / rider / day", 33, NUM0)
RFIN  = 22; label_calc(s, RFIN, "Extra orders/day to pay for rider",
                       f"=B{RCOST}/{D_CMO}/30", NUM1, "financial breakeven (trivial)")
RCAP  = 23; label_calc(s, RCAP, "2-rider delivery capacity / day", f"=2*B{RTRIP}", NUM0)
RTRIG = 24; label_calc(s, RTRIG, "Recommended hire trigger (orders/day)",
                       f"=0.85*B{RCAP}", NUM0, "hire before fleet saturates (SLA)", key=True)

# ---- Q4: discount impact ----
banner(s, 26, "Q4  —  DISCOUNT IMPACT (5-pizza order)")
Q4QTY = 27; label_input(s, Q4QTY, "Order quantity", 5, NUM0)
Q4FOOD= 28; label_calc(s, Q4FOOD, "Food cost / pizza (Cheese Burst)",
                       f"={CB_TOT}-{F_PKG}-{F_DELV}", RUPEE, "from COGS sheet, ex pkg/delivery")
Q4UNIT= 29; label_calc(s, Q4UNIT, "Unit selling price", f"='Bill Calculator'!B11", RUPEE)
Q4SUB = 30; label_calc(s, Q4SUB, "Subtotal (ex-GST)", f"=B{Q4UNIT}*B{Q4QTY}", RUPEE)
Q4DISC= 31; label_calc(s, Q4DISC, "Discount (10%)", f"=B{Q4SUB}*{F_DPCT}", RUPEE)
Q4VAR = 32; label_calc(s, Q4VAR, "Variable cost (food x qty + pkg x qty + 1 delivery)",
                       f"=B{Q4FOOD}*B{Q4QTY}+{F_PKG}*B{Q4QTY}+{F_DELV}", RUPEE)
Q4CMN = 33; label_calc(s, Q4CMN, "CM without discount", f"=B{Q4SUB}-B{Q4VAR}", RUPEE)
Q4CMW = 34; label_calc(s, Q4CMW, "CM with discount", f"=B{Q4SUB}-B{Q4DISC}-B{Q4VAR}", RUPEE)
Q4HIT = 35; label_calc(s, Q4HIT, "Contribution given up by discount",
                       f"=B{Q4CMN}-B{Q4CMW}", RUPEE, key=True)

# ---- Q5: ingredient inflation ----
banner(s, 37, "Q5  —  COGS INFLATION")
INFL  = 38; label_input(s, INFL, "Ingredient inflation %", 0.12, PCT)
NING  = 39; label_calc(s, NING, "New ingredient COGS / order", f"={F_ING}*(1+B{INFL})", RUPEE)
NCM   = 40; label_calc(s, NCM, "New CM / order",
                       f"={D_BAOV}-B{NING}-{F_PKG}-{F_DELV}-{D_BAOV}*{F_GW}", RUPEE)
NORD  = 41; label_calc(s, NORD, "Orders/month to hold EBITDA",
                       f"=({D_EBIT}+{F_FIXED})/B{NCM}", NUM0)
ADDD  = 42; label_calc(s, ADDD, "Extra orders / day needed",
                       f"=(B{NORD}-{D_MORD})/30", NUM1, key=True)

# ---- Q6: BI metrics (notes) ----
banner(s, 44, "Q6  —  BI METRICS FROM ORDER LOG")
notes = [
    "1. Basket / AOV analysis  ->  upsell & menu engineering (revenue drops to profit)",
    "2. Demand by hour x day-of-week  ->  staff & rider scheduling (cut idle cost, hold SLA)",
    "3. Item-level contribution-margin ranking  ->  promote high-margin, retire low-margin",
    "Bonus: discount-effectiveness (order-size distribution) + repeat-customer rate",
]
for k, t in enumerate(notes):
    s.cell(row=45 + k, column=1, value=t)

# ---- legend on Inputs ----
ws.cell(row=INVEST + 3, column=1, value="LEGEND:").font = BOLD
leg = ws.cell(row=INVEST + 4, column=1, value="Yellow = editable input")
leg.fill = INPUT_FILL
ws.cell(row=INVEST + 5, column=1, value="Green = key result   |   Blue = computed").fill = CALC_FILL

wb.save("M:/mouli-projects/pizzaflow/SliceMatic_Calculator.xlsx")
print("Workbook written: SliceMatic_Calculator.xlsx")
